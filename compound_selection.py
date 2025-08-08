#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 30 16:05:00 2024

@author: guohan
"""

import os, sys
from pathlib import Path
import pandas as pd
import numpy as np
from collections import Counter

from rdkit import Chem
from rdkit.Chem import rdFMCS
from rdkit.Chem.Draw import rdMolDraw2D, MolsToGridImage
from chembl_structure_pipeline import *
from chembl_structure_pipeline.checker import *
from rdkit.Chem.Draw import rdMolDraw2D

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XSImage
from openpyxl.styles import Alignment


from utils.tools import remove_unnamed_columns


### Get cluster labels ###
def get_clusterLabel(input_file_SMILES, input_file_clusterLabel, **kwargs):
    """
    Get cluster labels.
    :param input_file_SMILES: str, path of the input SMILES file
    :param input_file_clusterLabel: str, path of the input cluster label file
    :param id_column_name_SMILESFile: str, name of the ID column in input_file_SMILES
    :param id_column_name_clusterLabelFile: str, name of the ID column in input_file_clusterLabel
    :param clusterLabel_column_name_clusterLabelFile: str, name of the cluster label column in input_file_clusterLabel
    :return: None
    """
    # files
    output_file = os.path.splitext(os.path.abspath(input_file_SMILES))[0] + '_ClusterLabel'

    # kwargs
    id_column_name_SMILESFile = kwargs.get('id_column_name_SMILESFile', 'ID')
    id_column_name_clusterLabelFile = kwargs.get('id_column_name_clusterLabelFile', 'ID')
    clusterLabel_column_name_clusterLabelFile = kwargs.get('clusterLabel_column_name_clusterLabelFile', 'MCS Cluster')

    # read files
    df_SMILES = pd.read_csv(input_file_SMILES)
    df_SMILES.rename(columns={id_column_name_SMILESFile:'ID'}, inplace=True)
    print('Number of rows in SMILES file:', df_SMILES.shape[0])
    df_clusterLabel = pd.read_csv(input_file_clusterLabel)
    df_clusterLabel.rename(columns={id_column_name_clusterLabelFile:'ID', clusterLabel_column_name_clusterLabelFile:'MCS_Cluster'}, inplace=True)
    df_clusterLabel = pd.DataFrame(df_clusterLabel, columns=['ID', 'MCS_Cluster'])
    print('Number of rows in cluster label file:', df_clusterLabel.shape[0])

    # merge
    df = pd.merge(df_SMILES, df_clusterLabel, how='inner', on=['ID'])
    df = pd.DataFrame(df, columns=df_SMILES.columns.tolist() + ['MCS_Cluster'])

    # write output file
    df = df.reset_index(drop=True)
    print('Number of rows in the output file:', df.shape[0])
    df = remove_unnamed_columns(df)
    df.to_csv(f'{output_file}_{df.shape[0]}.csv')
    print('Getting cluster label is done.')


### Select representatives ###
def select_cmpds_from_clusters(input_file, clusterLabel_column_name, selection_method, **kwargs):
    """
    Select compounds from each cluster.
    :param input_file: str, path of the input file.
    :param clusterLabel_column_name: str, name of the cluster label column.
    :param selection_method: str, method to select representatives, allowed values include 'best', 'smallest', 'centroid_by_MCS'.

    :param outlier_label: label of outliers.
    :param keep_outlier: bool, specify whether to keep outliers.
    :param count_per_cluster: int, larger than 1, request count for each cluster.
    :param percentage_per_cluster: float, less than 1 and larger than 0, request percentage of compounds for each cluster.

    :param id_column_name: str, name of the ID column.
    :param SMILES_column_name: str, name of the SMILES column.
    :param dockingScore_column_name: str, name of the docking score column.
    :param property_rules: dict, dictionary of property rules, each key-value is defined as column_name:function.
    :param min_num_rules: int, minimum number of rules a compound need to satisfy.

    :param addCentroid: bool, specify whether to add centroid.

    :return: None
    """
    # files
    output_file = os.path.splitext(os.path.abspath(input_file))[0]+'_Representative'

    # read input file and define DataFrame for representatives
    df = pd.read_csv(input_file)
    df_representative = None

    # get cluster labels and outlier label
    clusterLabel_set = set(df[clusterLabel_column_name])
    outlier_label = kwargs.get('outlier_label', 0)
    keep_outlier = kwargs.get('keep_outlier', True)

    addCentroid = kwargs.get('addCentroid', False)

    # process clusters
    for label in clusterLabel_set:
        print(f'Cluster label={label}...')
        df_cluster = df[df[clusterLabel_column_name] == label]
        # get request count for outliers
        if label == outlier_label:
            if keep_outlier:
                count = df_cluster.shape[0]   # keep outliers
            else:
                continue
        # get request count for other clusters
        else:
            if 'count_per_cluster' in kwargs:
                count = kwargs.get('count_per_cluster')
                print(f'Request count number is {count}.')
            elif 'percentage_per_cluster' in kwargs:
                percentage = kwargs.get('percentage_per_cluster')
                count = round(df_cluster.shape[0] * percentage)
                if keep_outlier:   # If keep_outlier == True, request at least one compound for each cluster
                    count = max(count, 1)
            elif selection_method in {'centroid_by_MCS'}:
                count = 1
            else:
                raise Exception('Error: Please define either count_per_cluster or percentage_per_cluster.')

        # calculate centroid, save to the MCS_Centroid column
        if addCentroid:
            SMILES_column_name = kwargs.get('SMILES_column_name', 'SMILES')
            SMILESs = df_cluster[SMILES_column_name].values.tolist()
            mcs_SMILES = cal_MCS_centroid(SMILESs)
            if label == outlier_label:
                df_cluster.insert(df_cluster.columns.get_loc(SMILES_column_name) + 1, 'MCS_Centroid', df_cluster[SMILES_column_name].tolist())
            else:
                df_cluster.insert(df_cluster.columns.get_loc(SMILES_column_name) + 1, 'MCS_Centroid', mcs_SMILES)

        ## get representatives from each cluster
        df_new_representative = pd.DataFrame()
        # get the best compounds (meet the most criteria and have the highest docking score)
        if selection_method == 'best':
            dockingScore_column_name = kwargs.get('dockingScore_column_name', 'Docking_Score')
            property_rules = kwargs.get('property_rules', {})
            min_num_rules = kwargs.get('min_num_rules', len(property_rules))
            df_new_representative = get_best_compound(df_cluster, count, dockingScore_column_name, property_rules, min_num_rules)
        # get the smallest compounds (the least heavy atoms)
        elif selection_method == 'smallest':
            SMILES_column_name = kwargs.get('SMILES_column_name', 'SMILES')
            dockingScore_column_name = kwargs.get('dockingScore_column_name', 'Docking_Score')
            df_new_representative = get_smallest_compound(df_cluster, count, SMILES_column_name, dockingScore_column_name)
        # get the MCS centroids
        elif selection_method == 'centroid_by_MCS':
            id_column_name = kwargs.get('id_column_name', 'ID')
            SMILES_column_name = kwargs.get('SMILES_column_name', 'SMILES')
            df_new_representative = get_centroid_by_mcs(df_cluster, label==outlier_label, id_column_name, SMILES_column_name,
                                                        clusterLabel_column_name)

        # concat representative df from each cluster
        if df_new_representative.shape[0] == 0:
            continue
        if df_representative is None:
            df_representative = df_new_representative
        else:
            df_representative = pd.concat([df_representative, df_new_representative], ignore_index=True, sort=False)

    # write output file
    df_representative = df_representative.reset_index(drop=True)
    print('Number of rows in the output file:', df_representative.shape[0])
    df_representative = remove_unnamed_columns(df_representative)
    df_representative.to_csv(f'{output_file}_{df_representative.shape[0]}.csv')
    print('Selection of representative compounds is done.')

    # create a .xlsx file, draw molecules with centroid highlighted
    if addCentroid:
        id_column_name = kwargs.get('id_column_name', 'ID')
        SMILES_column_name = kwargs.get('SMILES_column_name', 'SMILES')
        gen_MCS_entroid_fig_column(df_representative, f'{output_file}_{df_representative.shape[0]}.xlsx', id_column_name,
                                   SMILES_column_name, 'MCS_Centroid')


def get_best_compound(df_cluster, count, dockingScore_column_name, property_rules, min_num_rules):
    """
    Helper function for select_cmpds_from_clusters. Select the best compounds from each cluster,
    which meet the most criteria and have the highest docking score.
    :param df_cluster: pd.DataFrame object which contains compounds from each cluster.
    :param count: int, request count for each cluster.
    :param dockingScore_column_name: str or None, name of the docking score column.
    :param property_rules: dict, dictionary of property rules, each key-value is defined as column_name:function.
    :param min_num_rules: int, minimum number of rules a compound need to satisfy.
    :return: pd.DataFrame object which contains selected representative compounds.
    """
    # columns
    COLUMNS = df_cluster.columns.tolist()

    # convert value to score (1 or 0) for each property
    for column, rule in property_rules.items():
        try:
            df_cluster.insert(df_cluster.shape[1], column+'_', [int(v) for v in df_cluster[column].apply(rule)])
        except Exception:
            print(f'Error: Rule for {column} column is not applied.')
            continue

    # compute compound property score (number of criteria met)
    property_column_names = property_rules.keys()
    property_score_column_names = [property+'_'for property in property_column_names]
    df_cluster.insert(df_cluster.shape[1], 'Property_Score', df_cluster[property_score_column_names].sum(axis=1).values.tolist())
    # get compounds that meet minimum number of criteria
    df_filtered = df_cluster[df_cluster['Property_Score'] >= min_num_rules]
    # sort compounds based on property score first then on docking score
    if dockingScore_column_name is not None:
        df_filtered = df_filtered.sort_values(by=['Property_Score', dockingScore_column_name], ascending=[False, False], ignore_index=True)
    else:
        df_filtered = df_filtered.sort_values(by=['Property_Score'], ascending=[False], ignore_index=True)
    # get representative compounds based on count
    n = min(count, df_filtered.shape[0])
    df_subset = df_filtered.loc[0:(n-1)]
    df_subset = pd.DataFrame(df_subset, columns=COLUMNS+['Property_Score'])

    return df_subset


def get_smallest_compound(df_cluster, count, SMILES_column_name, dockingScore_column_name):
    """
    Helper function for select_cmpds_from_clusters. Select the smallest compounds.
    (i.e., the least heavy atoms) from each cluster.
    :param df_cluster: pd.DataFrame object which contains compounds from each cluster.
    :param count: int, request count for each cluster.
    :param SMILES_column_name: str, name of the SMILES column.
    :param dockingScore_column_name: str, name of the docking score column.
    :return: pd.DataFrame object which contains selected representative compounds.
    """
    # columns
    COLUMNS = df_cluster.columns.tolist()

    # compute the number of heavy atoms
    NHA = [Chem.MolFromSmiles(smiles).GetNumHeavyAtoms() for smiles in df_cluster[SMILES_column_name].values.tolist()]
    df_cluster.insert(len(COLUMNS), 'NHA', NHA)
    # sort compounds based on number of heavy atoms first then on docking score
    df_cluster = df_cluster.sort_values(by=['NHA', dockingScore_column_name], ascending=[True, True], ignore_index=True)
    # get representatives compounds based on count
    n = min(count, df_cluster.shape[0])
    df_subset = df_cluster.loc[0:(n-1)]
    df_subset = pd.DataFrame(df_subset, columns=COLUMNS)

    return df_subset


def get_centroid_by_mcs(df_cluster, is_outlier, id_column_name, SMILES_column_name, clusterLabel_column_name):
    """
    Helper function for select_cmpds_from_clusters. Select the centroid defined as the MCS of the cluster.
    :param df_cluster: pd.DataFrame object which contains compounds from each cluster.
    :param is_outlier: bool, specify whether this cluster is outliers.
    :param id_column_name: str, name of the ID column.
    :param SMILES_column_name: str, name of the SMILES column.
    :param clusterLabel_column_name: str, name of the cluster label column.
    :return: pd.DataFrame object which contains selected representative compounds.
    """
    # return outliers directly
    if is_outlier:
        return df_cluster

    # compute MCS
    SMILESs = df_cluster[SMILES_column_name].values.tolist()
    mcs_SMILES = cal_MCS_centroid(SMILESs)

    # centroid DataFrame
    label = df_cluster[clusterLabel_column_name].tolist()[0]
    df_centroid = pd.DataFrame({id_column_name:[f'cluster{label}_MCS_centroid'], SMILES_column_name:[mcs_SMILES],
                                clusterLabel_column_name:[label]})

    return df_centroid


### Add centroid figure column ###
def add_centroid_figure_column(input_file, id_column_name, SMILES_column_name, centroid_column_name = 'MCS_Centroid'):
    """
    Add centroid figure column to the centroid column
    :param input_file: str, path of the input file.
    :param id_column_name: str, name of the ID column.
    :param SMILES_column_name: str, name of the SMILES column.
    :param MCS_centroid_column_name: str, name of the centroid column.
    :return: None
    """
    output_file = os.path.splitext(os.path.abspath(input_file))[0]+'.xlsx'
    df = pd.read_csv(input_file)
    gen_MCS_entroid_fig_column(df, output_file, id_column_name, SMILES_column_name, centroid_column_name)


### Helper functions to calculate centroid and plot centroid ###
def cal_MCS_centroid(SMILESs):
    """
    Helper function to calculate the centroid defined as the MCS of the cluster.
    :param SMILESs: list of str, list of SMILES.
    :return: str, the MCS centroid SMILES.
    """
    if len(SMILESs) == 0:
        return None
    elif len(SMILESs) == 1:
        return SMILESs[0]

    # compute MCS
    mol_list = [Chem.MolFromSmiles(smiles) for smiles in SMILESs]
    mcs_SMARTS = rdFMCS.FindMCS(mol_list, ringMatchesRingOnly=True, completeRingsOnly=True, bondCompare=rdFMCS.BondCompare.CompareOrderExact).smartsString

    return mcs_SMARTS


def draw_molecule_with_centroid(ID, SMILES, centroid):
    """
    Helper function to draw molecule with the centroid highlighted.
    :param ID: str, ID of the compound.
    :param SMILES: str, compound SMILES.
    :param centroid: str, centroid SMARTS.
    :return: str, path of the image file in .png format.
    """
    # get matched atoms and bonds
    mol = Chem.MolFromSmiles(SMILES)
    centroid_mol = Chem.MolFromSmarts(centroid)
    match_atoms = list(mol.GetSubstructMatch(centroid_mol))
    match_bonds = []
    for bond in centroid_mol.GetBonds():
        atom1 = match_atoms[bond.GetBeginAtomIdx()]
        atom2 = match_atoms[bond.GetEndAtomIdx()]
        match_bonds.append(mol.GetBondBetweenAtoms(atom1, atom2).GetIdx())

    # plot and save img
    d = rdMolDraw2D.MolDraw2DCairo(250, 250)
    rdMolDraw2D.PrepareAndDrawMolecule(d, mol, highlightAtoms=match_atoms, highlightBonds=match_bonds)
    d.FinishDrawing()
    file_path = Path(f'temp/{ID}.png')
    if not file_path.parent.exists():
        file_path.parent.mkdir()
    d.WriteDrawingText(f'temp/{ID}.png')


def gen_MCS_entroid_fig_column(df, output_file, id_column_name, SMILES_column_name, MCS_centroid_column_name = 'MCS_Centroid'):
    """
    Helper function, generate MCS centroid figure column
    :param df: pd.DataFrame object.
    :param id_column_name: str, name of the ID column.
    :param SMILES_column_name: str, name of the SMILES column.
    :param MCS_centroid_column_name: str, name of the MCS centroid column.
    :return:
    """
    # draw molecules in SMILES_column_name with MCS highlighted
    df.apply(lambda row: draw_molecule_with_centroid(row[id_column_name], row[SMILES_column_name], row[MCS_centroid_column_name]), axis=1)
    df.loc[:, MCS_centroid_column_name] = ''

    # create .xlsx file, add content to excel
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.append(df.columns.tolist())
    for i, row in df.iterrows():
        ws.append(row.tolist())

    # initialize the molecular figure column
    num_row = df.shape[0]
    MCS_centroid_column_number = df.columns.get_loc(MCS_centroid_column_name) + 1
    MCS_centroid_column_letter = get_column_letter(MCS_centroid_column_number)
    # ws[img_column_letter + str(1)] = 'Molecular_Figure'
    index_start, index_end = 2, 2 + num_row

    # resize the cell
    WIDTH_FUDGE = 1080 / 1920 * (1 - 72 / 95.25) # 72 is a standart dpi; if the screen resolution is 1920*1080
    HEIGHT_FUDGE = 3 / 4
    img_width, img_height = 250, 250
    ws.column_dimensions[MCS_centroid_column_letter].width = img_width*WIDTH_FUDGE

    # add image
    for index, ID in enumerate(df[id_column_name].tolist()):
        ws.row_dimensions[index_start + index].height = img_height * HEIGHT_FUDGE
        img = XSImage(f'temp/{ID}.png')
        ws.add_image(img, MCS_centroid_column_letter + str(index_start + index))

    # alignment
    al = Alignment(horizontal='general', vertical='top')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = al

    # write output file
    wb.save(output_file)




if __name__ == '__main__':
    ### Get cluster labels ###
    # input_file_SMILES = 'tests/test_SMILES_file.csv'
    # input_file_clusterLabel = 'tests/test_clusterLabel_file.csv'
    # get_clusterLabel(input_file_SMILES, input_file_clusterLabel,
    #                  id_column_name_SMILESFile='ID', id_column_name_clusterLabelFile='ID',
    #                  clusterLabel_column_name_clusterLabelFile='MCS Cluster 0.7')


    ### Select representatives ###
    # print(pd.options.mode.copy_on_write)
    input_file = 'tests/test_select_representatives.csv'
    clusterLabel_column_name = 'MCS_Cluster'
    # Select the best compounds #
    property_rules = {'MW': lambda mw: mw <= 500, 'logP': lambda logp: logp > 0.0, 'HBD': lambda hbd: hbd <= 3}
    select_cmpds_from_clusters(input_file, clusterLabel_column_name, selection_method='best',
                               outlier_label=0, keep_outlier=True, count_per_cluster=2,
                               SMILES_column_name='Cleaned_SMILES', dockingScore_column_name='Docking_Score',
                               property_rules=property_rules, min_num_rules=2, addCentroid=True)
    # Select the smallest compound #
    # select_cmpds_from_clusters(input_file, clusterLabel_column_name, selection_method='smallest',
    #                            outlier_label=0, keep_outlier=True, count_per_cluster=1,
    #                            SMILES_column_name='Cleaned_SMILES', dockingScore_column_name='Docking_Score',
    #                            addCentroid=True)
    # Get the centroid defined as MCS #
    # select_cmpds_from_clusters(input_file, clusterLabel_column_name, selection_method='centroid_by_MCS',
    #                            outlier_label=0, keep_outlier=True,
    #                            id_column_name = 'ID', SMILES_column_name='Cleaned_SMILES')


    ### Add centroid figure column ###
    # input_file = 'tests/test_add_centroid_figure_column.csv'
    # add_centroid_figure_column(input_file, id_column_name = 'ID', SMILES_column_name = 'Cleaned_SMILES', centroid_column_name='MCS_Centroid')


    # plotting
    # ID = 'cmpd_216'
    # SMILES = 'N=c1nc(N(Cc2ccc(C(F)(F)F)cc2)c2ccc(S(N)(=O)=O)cc2)cc[nH]1'
    # centroid = 'N=C1:N:C:C:C(N(Cc2ccc(C(F)(F)F)cc2)c2ccc(S(N)(=O)=O)cc2):N:1'
    # centroid = 'N=C1:[nH]:C:C:C(N(Cc2ccc(C(F)(F)F)cc2)c2ccc(S(N)(=O)=O)cc2):N:1'
    # draw_molecule_with_centroid(ID, SMILES, centroid)


